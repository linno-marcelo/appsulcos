import os
import openpyxl
from kivy.lang import Builder
from kivymd.app import MDApp
from kivy.utils import platform
from android.storage import app_storage_path
from android.permissions import request_permissions, Permission


class MyApp(MDApp):
    def build(self):
        self.title = "Formulário de Profundidade de Sulco"
        return Builder.load_file("main.kv")

    def salvar_informacoes(self):
        dados = {
            "equipamento": self.root.ids.equipamento.text,
            "cs": self.root.ids.cs.text,
            "data": self.root.ids.data.text,
            "codigo_pneu": self.root.ids.codigo_pneu.text,
            "sulco1": self.root.ids.sulco1.text,
            "sulco2": self.root.ids.sulco2.text,
            "sulco3": self.root.ids.sulco3.text,
            "medida": self.root.ids.medida.text,
            "calibrada": self.root.ids.calibrada.text
        }

        # Define the directory and file path
        if platform == 'android':
            from android.storage import primary_external_storage_path
            dir_path = os.path.join(primary_external_storage_path(), 'dados')
            if not os.path.exists(dir_path):
                os.makedirs(dir_path)
            file_path = os.path.join(dir_path, 'dados.xlsx')
        else:
            dir_path = os.path.join(os.getcwd(), 'dados')
            if not os.path.exists(dir_path):
                os.makedirs(dir_path)
            file_path = os.path.join(dir_path, 'dados.xlsx')

        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            if ws.max_row == 1:
                ws.append(["Equipamento", "CS", "Data", "Código Pneu",
                          "Sulco1", "Sulco2", "Sulco3", "Medida", "Calibrada"])
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Equipamento", "CS", "Data", "Código Pneu",
                      "Sulco1", "Sulco2", "Sulco3", "Medida", "Calibrada"])

        ws.append([dados["equipamento"], dados["cs"], dados["data"], dados["codigo_pneu"],
                   dados["sulco1"], dados["sulco2"], dados["sulco3"], dados["medida"], dados["calibrada"]])
        wb.save(file_path)
        print("Informações salvas com sucesso!")

    def on_start(self):
        if platform == 'android':
            request_permissions(
                [Permission.WRITE_EXTERNAL_STORAGE, Permission.READ_EXTERNAL_STORAGE])


MyApp().run()
